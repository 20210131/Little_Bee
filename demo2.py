import datetime
import time
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
#字节跳洞1.0
#用于我偷懒记录数据！！！
# 定义保存点击状态的文件路径
CLICK_STATUS_FILE = "click_status3.txt"

def login(username, password):
    # 初始化浏览器
    driver = webdriver.Chrome()  # 也可以使用其他浏览器的驱动

    try:
        # 打开网页
        driver.get("https://btdjhg.beyondh.com:8101/")

        # 输入账号
        username_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "idUserNameInput")))
        username_input.send_keys(username)

        # 等待2秒，等到密码输入框出现
        time.sleep(2)

        # 输入密码
        password_input = driver.find_element(By.ID, "idPasswordInput")
        password_input.send_keys(password)

        # 点击登录按钮
        login_button = driver.find_element(By.ID, "idLoginButton")
        login_button.click()

        # 等待3秒，等待选择门店出现
        time.sleep(3)

        return driver

    except Exception as e:
        print("登录失败:", e)
        driver.quit()
        return None

def select_store(driver, store_xpath):
    try:
        # 点击选择门店
        select_store = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='f_l']/div[1]")))
        select_store.click()

        # 等待门店选项出现
        time.sleep(3)

        # 点击对应门店
        store_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, store_xpath)))
        store_name = store_element.text.strip()
        print(f"===== 正在爬取门店 '{store_name}' 数据 =====")
        store_element.click()

        # 等待门店选择成功
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@class='c_hotel_footer']/button[2]")))

        # 点击确定门店按钮
        confirm_store_button = driver.find_element(By.XPATH, "//div[@class='c_hotel_footer']/button[2]")
        confirm_store_button.click()

        # 等待页面加载
        time.sleep(5)

        return store_name

    except Exception as e:
        print("选择门店失败:", e)
        return None

def extract_data(driver, store_name):
    try:
        # 检查是否已经执行过一次操作
        already_clicked_more_button = check_click_status()

        # 查看是否有下拉菜单，并且还没有执行过操作
        try:
            more_button = driver.find_element(By.XPATH, "//div[@class='more_btn']/span[1]")
            if more_button.is_displayed() and not already_clicked_more_button:
                more_button.click()
                time.sleep(2)
                already_clicked_more_button = True
                save_click_status(already_clicked_more_button)
        except:
            pass

        
       # 获取数据
        room_types = driver.find_elements(By.XPATH, "//DIV[@class='roomTypeFilter']")
        
        # 获取当前时间，以小时为单位
        current_hour = datetime.datetime.now().hour
        
        # 创建或打开对应的Excel文件
        excel_filename = f"{store_name}_data.xlsx"
        try:
            wb = openpyxl.load_workbook(excel_filename)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        # 创建新的Sheet，以当前时间命名
        sheet_name = str(current_hour)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        
        # 将数据写入Sheet
        for room_type in room_types:
            data = room_type.text.strip()
            ws.append([data])
        
        # 保存Excel文件
        wb.save(excel_filename)
        
        print(f"数据已写入 {excel_filename} 的 {sheet_name} 工作表")

    except Exception as e:
        print("提取数据失败:", e)

def check_click_status():
    # 检查是否已经执行过点击操作
    if os.path.exists(CLICK_STATUS_FILE):
        with open(CLICK_STATUS_FILE, "r") as file:
            status = file.read().strip().lower()
            return status == "clicked"
    return False

def save_click_status(clicked):
    # 保存点击状态到文件
    with open(CLICK_STATUS_FILE, "w") as file:
        file.write("clicked" if clicked else "not clicked")

if __name__ == "__main__":
    # 输入您的账号和密码
    username_input = "0000"
    password_input = "abc1234."

    # 登录并获取driver对象
    driver = login(username_input, password_input)

    if driver:
        try:
            # 选择门店并提取数据
            selected_store_xpaths = [
                "//div[@class='ant-tree-list-holder-inner']/div[2]/span[3]/span[1]", # 红谷滩店
                "//div[@class='ant-tree-list-holder-inner']/div[3]/span[3]/span[1]", # 新余店
                "//div[@class='ant-tree-list-holder-inner']/div[5]/span[3]/span[1]", # 太原追风店
                "//div[@class='ant-tree-list-holder-inner']/div[7]/span[3]/span[1]", # 南大店
                "//div[@class='ant-tree-list-holder-inner']/div[9]/span[3]/span[1]", # 绳金塔店
                "//div[@class='ant-tree-list-holder-inner']/div[8]/span[3]/span[1]"  # 宜春追风店
            ]

            for store_xpath in selected_store_xpaths:
                store_name = select_store(driver, store_xpath)
                if store_name:
                    extract_data(driver, store_name)

        finally:
            # 关闭浏览器
            driver.quit()
    else:
        print("登录失败，请检查账号和密码。")
file = open("click_status3.txt", 'w').close()